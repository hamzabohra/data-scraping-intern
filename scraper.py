import json
import re
import asyncio
import os
import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from crawl4ai import AsyncWebCrawler, CrawlerRunConfig, CacheMode, BrowserConfig
import openai
from dotenv import load_dotenv

load_dotenv()
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY")

if not OPENROUTER_API_KEY:
    print("❌ ERROR: OPENROUTER_API_KEY not found in .env file.")
    exit(1)

client = openai.OpenAI(
    base_url="https://openrouter.ai/api/v1",
    api_key=OPENROUTER_API_KEY,
)

OUTPUT_FILE = "university_courses_cd.xlsx"
CD_HUB = "https://collegedunia.com/india-colleges"

def clean(text):
    if not text or str(text).lower() in ["n/a", "none", "unknown", "varies", "null", "vairies"]:
        return "Not specified"
    return " ".join(str(text).replace("\n", " ").split())

async def get_top_universities(crawler):
    """Dynamically get exactly 5 top universities from Collegedunia."""
    print("🔍 [DISCOVERY] Fetching Top Universities from Collegedunia Aggregator...")
    # magic=True helps bypass basic bot protections on aggregator pages
    res = await crawler.arun(url=CD_HUB, config=CrawlerRunConfig(cache_mode=CacheMode.BYPASS, magic=True))
    
    universities = []
    seen = set()
    
    if res.success and res.links.get('internal'):
        for l in res.links['internal']:
            href = l['href']
            # We want only the base university url: https://collegedunia.com/university/ID-NAME
            match = re.search(r'(https://collegedunia\.com/university/\d+-[^/]+)', href)
            if match:
                url_clean = match.group(1)
                text = str(l.get('text', '')).strip()
                if len(text) > 3 and not text.startswith('http'):
                    if url_clean not in seen:
                        seen.add(url_clean)
                        universities.append({
                            'name': text,
                            'url': url_clean,
                        })
            if len(universities) >= 5:
                break
                
    # Fallback to AI extraction if the internal link parsing missed something
    if len(universities) < 5:
        print("    ↳ [!] Using AI DOM parsing fallback for CD University Discovery...")
        prompt = f"""Extract exactly 5 unique Indian universities from this ranking page content.
Return ONLY a JSON array of objects with keys: "name" (University Name), "url" (The absolute URL to the university profile on collegedunia.com).
CONTENT: {res.markdown[:15000]}"""
        try:
            resp = client.chat.completions.create(
                model="stepfun/step-3.5-flash:free",
                messages=[{"role": "user", "content": prompt}]
            )
            json_str = re.search(r'\[.*\]', resp.choices[0].message.content, re.DOTALL)
            if json_str:
                universities = json.loads(json_str.group(0))[:5]
        except Exception as e:
            print(f"      [!] Extraction error: {e}")
            
    print(f"  ↳ Found {len(universities)} universities.")
    for u in universities:
        print(f"    - {u['name']}")
        
    return universities[:5]

async def get_university_metadata(raw_name, uni_url):
    """Extract official website details using LLM."""
    print(f"\n  ↳ [METADATA] Identifying details for {raw_name}...")
    prompt = f"""Provide the official metadata for the university found at this URL: {uni_url}. 
The provided raw name might contain junk data like fees or 'PG Program'. Fix it.
Return ONLY a valid JSON object. Do not wrap in markdown blocks.
Required keys:
"official_name": The clean, official name of the university (e.g., "IIT Bombay", "IIM Ahmedabad")
"country": 2-letter country code (e.g., "IN" for India)
"city": City where the main campus is located
"website": The official website URL of the university (e.g. https://www.iitb.ac.in. MUST be the official university domain, NOT collegedunia or shiksha)
"""
    retries = 3
    for attempt in range(retries):
        try:
            resp = client.chat.completions.create(
                model="stepfun/step-3.5-flash:free",
                messages=[{"role": "user", "content": prompt}],
            )
            text = resp.choices[0].message.content
            match = re.search(r'\{.*\}', text, re.DOTALL)
            if match:
                data = json.loads(match.group(0))
                return {
                    "name": clean(data.get("official_name", raw_name)),
                    "country": data.get("country", "IN"),
                    "city": data.get("city", "Unknown City"),
                    "website": data.get("website", "")
                }
            return {"name": clean(raw_name), "country": "IN", "city": "Unknown", "website": ""}
        except Exception as e:
            if "429" in str(e) and attempt < retries - 1:
                print(f"      [!] Rate limited. Sleeping 5s then retrying...")
                await asyncio.sleep(5)
                continue
            print(f"      [!] Metadata Extraction error: {e}")
            return {"name": clean(raw_name), "country": "IN", "city": "Unknown", "website": ""}

def find_full_time(obj):
    """Recursively search for the 'full_time' courses array in Collegedunia's __NEXT_DATA__ JSON"""
    if isinstance(obj, dict):
        if 'full_time' in obj and isinstance(obj['full_time'], list) and len(obj['full_time']) > 0 and isinstance(obj['full_time'][0], dict) and 'stream' in obj['full_time'][0]:
            return obj['full_time']
        for v in obj.values():
            r = find_full_time(v)
            if r: return r
    elif isinstance(obj, list):
        for item in obj:
            r = find_full_time(item)
            if r: return r
    return None

async def extract_course_specialization_slugs(crawler, uni_base_url, uni_name):
    """Fetch the college page, extract __NEXT_DATA__ JSON, and find all specialization JSON objects."""
    print(f"\n  ↳ [METADATA] Discovering explicit course URLs for {uni_name}...")
    
    if "courses-fees" in uni_base_url:
        courses_fees_url = uni_base_url
    else:
        courses_fees_url = uni_base_url.rstrip('/') + '/courses-fees'
    print(f"      [DEBUG] Fetching URL: {courses_fees_url}")
    # We fetch the courses-fees page since it contains the full JSON payload
    res = await crawler.arun(url=courses_fees_url, config=CrawlerRunConfig(cache_mode=CacheMode.BYPASS, magic=True))
    if not res.success or not res.html:
        print("        [!] Failed to load college page HTML.")
        return []

    soup = BeautifulSoup(res.html, 'html.parser')
    next_data = soup.find('script', id='__NEXT_DATA__')
    print(f"      [DEBUG] Searched for __NEXT_DATA__, Found: {bool(next_data)}")

    if not next_data:
        print(f"      [!] No __NEXT_DATA__ payload found for {uni_name}")
        print(f"      [DEBUG] Page Title: {soup.title.text if soup.title else 'No Title'}")
        await asyncio.sleep(2) # Add sleep as per instruction
        return []
    
    specializations = []
    
    if next_data:
        try:
            print(f"      [DEBUG] Next Data length: {len(next_data.string)}")
            data = json.loads(next_data.string)
            print(f"      [DEBUG] Top level keys: {list(data.keys())}")
            
            try:
                pageProps = data.get("props", {}).get("pageProps", {})
                print(f"      [DEBUG] pageProps keys: {list(pageProps.keys())}")
            except Exception as e:
                 print(f"      [DEBUG] Error getting pageProps: {e}")
                 
            full_time_groups = find_full_time(data)
            print(f"      [DEBUG] full_time_groups found: {bool(full_time_groups)}")
            if full_time_groups:
                for group in full_time_groups:
                    group_name = group.get('course_tag_name', 'Unknown')
                    streams = group.get('stream', [])
                    for s in streams:
                        slug = s.get('sub_course_slug')
                        url_part = s.get('url')
                        
                        actual_slug = slug
                        if not actual_slug and url_part:
                            actual_slug = url_part
                            
                        if actual_slug:
                            specializations.append({
                                'main_group': group_name,
                                'slug': actual_slug,
                                'course_name': s.get('course_name') or s.get('display_course_name') or actual_slug
                            })
                print(f"      ↳ Extracted {len(specializations)} sub-course slugs from Next.js payload.")
            else:
                print(f"      [DEBUG] full_time_groups is None! Top-level keys: {list(data.keys())}")
                if 'props' in data and 'pageProps' in data['props']:
                    print(f"      [DEBUG] pageProps keys: {list(data['props']['pageProps'].keys())}")
                await asyncio.sleep(2) # Add sleep as per instruction
        except Exception as e:
            print(f"      [!] Failed to parse __NEXT_DATA__ JSON: {e}")
    else:
        print("      [!] Could not find __NEXT_DATA__ script block.")
        
    return specializations

async def extract_course_details(crawler, base_url, spec):
    """Scrape a specific specialization page (e.g., M.Tech CSE) for its metadata table."""
    spec_url = base_url.rstrip('/') + '/' + spec['slug']
    print(f"      🤖 Extracting specific course details: {spec['course_name']} from {spec_url}")
    
    res = await crawler.arun(url=spec_url, config=CrawlerRunConfig(cache_mode=CacheMode.BYPASS, magic=True))
    if not res.success or not res.markdown:
        print("        [!] Failed to load specialization page.")
        return None

    prompt = f"""
You are an expert data scraper.
Extract the "Duration", "Fees", "Eligibility", "Level", and "Discipline" for the specific course from the following markdown content of a Collegedunia course specialization page.
The text contains a "Course Highlights" table and an "Eligibility" or "Fee Structure" section. Look carefully.

Extract these fields into a strict JSON object:
- "duration": Course length (e.g., "2 Years", "4 Years"). Search the entire text for words like "Duration", "Course Period", "Years".
- "fees": Total tuition/course fees (e.g., "INR 1.24 Lakhs", "80,000"). Search the entire text for numbers relating to fees.
- "eligibility": Primary entry requirement. Search the entire text for "Eligibility", "Accepted Exams", "Selection Criteria".
- "level": The academic level. ONLY use one of: "Undergraduate", "Postgraduate", "Doctorate", "Diploma", "Certificate". Infer from context (e.g. B.Tech/B.Sc is Undergraduate, M.Tech/M.Sc is Postgraduate, Ph.D is Doctorate).
- "discipline": The broad field of study (e.g. "Engineering", "Medical", "Management", "Science", "Law", "Arts"). Infer from context.

If you cannot find the EXACT value, do your best to infer it from the context instead of giving up. Only use "Not specified" if absolutely necessary.

Return ONLY a valid JSON object. Do not wrap in markdown blocks like ```json.
CONTENT:
{res.markdown[:40000]}
"""
    retries = 4
    for attempt in range(retries):
        try:
            resp = client.chat.completions.create(
                model="stepfun/step-3.5-flash:free",
                messages=[{"role": "user", "content": prompt}]
            )
            
            text = resp.choices[0].message.content
            match = re.search(r'\{.*\}', text, re.DOTALL)
            if match:
                data = json.loads(match.group(0))
                
                # Use LLM extracted discipline, but fallback to Next.js tag if LLM fails
                extracted_discipline = clean(data.get("discipline"))
                if extracted_discipline.lower() in ["not specified", "unknown", ""]:
                    extracted_discipline = clean(spec['main_group'])
                    
                return {
                    "name": clean(spec['course_name']),
                    "level": clean(data.get("level", "Not specified")),
                    "discipline": extracted_discipline,
                    "duration": clean(data.get("duration")),
                    "fees": clean(data.get("fees")),
                    "eligibility": clean(data.get("eligibility")),
                    "url": spec_url
                }
            else:
                return None
        except Exception as e:
            if "429" in str(e) and attempt < retries - 1:
                print(f"      [!] Rate limited. Sleeping 6s then retrying...")
                await asyncio.sleep(6)
                continue
            print(f"      [!] Course Extraction error: {e}")
            return None

def save_excel(u_rows, c_rows):
    """Generates the 2-sheet relational Excel file."""
    print(f"\n💾 Generating professional Excel: {OUTPUT_FILE}")
    wb = openpyxl.Workbook()
    
    ws1 = wb.active
    ws1.title = "Universities"
    ws1.append(["University ID", "University Name", "Country", "City", "Website"])
    for u in u_rows:
        ws1.append([u['id'], u['name'], u['country'], u['city'], u['website']])
    
    ws2 = wb.create_sheet("Courses")
    ws2.append(["Course ID", "University ID", "Course Name", "Level", "Discipline", "Duration", "Total Tuition Fees", "Eligibility", "URL"])
    for c in c_rows:
        ws2.append([c['id'], c['u_id'], c['name'], c['level'], c['discipline'], c['duration'], c['fees'], c['eligibility'], c['url']])
    
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 25
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(OUTPUT_FILE)
    print("✅ Excel file saved successfully!")

async def main():
    print("=" * 70)
    print("  🚀 COLLEGEDUNIA-STYLE ULTRA-EXTRACTOR v7")
    print("=" * 70)
    
    uni_rows, course_rows = [], []
    c_idx = 1
    
    browser_config = BrowserConfig(headless=True)
    
    async with AsyncWebCrawler(config=browser_config) as crawler:
        # Step 1: Discover top universities from Collegedunia aggregator
        target_unis = await get_top_universities(crawler)
        
        for u_idx, uni in enumerate(target_unis, 1):
            raw_name = uni['name']
            uni_url = uni['url']
            
            # Hotfix: Collegedunia's aggregator serves broken ghost listings for some universities.
            # These correct IDs were verified by browsing collegedunia.com directly.
            URL_HOTFIXES = {
                '25494-iima-indian-institute-of-management-ahmedabad': 'https://collegedunia.com/university/25700-indian-institute-of-management-iima-ahmedabad-courses-fees',
                '1534-aiims-new-delhi': 'https://collegedunia.com/university/25446-all-india-institute-of-medical-sciences-aiims-new-delhi',
                '3294-s-aiims-new-delhi': 'https://collegedunia.com/university/25446-all-india-institute-of-medical-sciences-aiims-new-delhi',
                'parul-university': 'https://collegedunia.com/university/55884-parul-university-vadodara',
            }
            for broken_key, fixed_url in URL_HOTFIXES.items():
                if broken_key in uni_url:
                    uni_url = fixed_url
                    break
                
            # The aggregator handles different root URLs. If it natively links to courses-fees, use it directly.
            if "courses-fees" in uni_url:
                courses_url = uni_url
            else:
                courses_url = uni_url.rstrip('/') + '/courses-fees'
                
            # Extract official university website and location
            meta = await get_university_metadata(raw_name, uni_url)
            clean_name = meta['name']
            
            uni_rows.append({
                "id": u_idx,
                "name": clean_name,
                "country": meta['country'],
                "city": meta['city'],
                "website": meta['website']
            })
            
            # Step 2: Extract specialization slugs from __NEXT_DATA__
            await asyncio.sleep(2) # Prevent Cloudflare ratelimit
            
            all_specs = await extract_course_specialization_slugs(crawler, courses_url, clean_name)
            
            if not all_specs:
                print("      [!] No explicit courses found to extract.")
                continue
                
            # Step 3: Fetch individual specialization pages for deep details
            valid_courses = 0
            for spec in all_specs:
                if valid_courses >= 5:
                    break
                details = await extract_course_details(crawler, uni_url, spec)
                if details:
                    course_rows.append({
                        "id": c_idx,
                        "u_id": u_idx,
                        "name": details.get("name"),
                        "level": details.get("level"),
                        "discipline": details.get("discipline"),
                        "duration": details.get("duration"),
                        "fees": details.get("fees"),
                        "eligibility": details.get("eligibility"),
                        "url": details.get("url")
                    })
                    c_idx += 1
                    valid_courses += 1
            
    # Save the consolidated Excel output
    save_excel(uni_rows, course_rows)

if __name__ == "__main__":
    asyncio.run(main())
